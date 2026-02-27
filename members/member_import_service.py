"""
Member Import Service
Following SOLID principles:
- SRP: Each class has a single responsibility
- OCP: Extensible for different file formats
- DIP: Depends on abstractions
"""

import csv
import io
import logging
from typing import Dict, List, Optional, Tuple
from decimal import Decimal
from datetime import datetime
from django.db import transaction
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from openpyxl import load_workbook

from .models import Member
from .utils import normalize_phone_number

logger = logging.getLogger(__name__)

# Maximum records per import to prevent timeouts
MAX_IMPORT_RECORDS = 5000
# Batch size for database operations
BATCH_SIZE = 100


class MemberImportValidator:
    """
    Validates individual member records.
    Following SRP: Only responsible for validation logic.
    """

    REQUIRED_FIELDS = ['first_name', 'last_name', 'phone_number']
    OPTIONAL_FIELDS = ['email', 'member_number']

    @staticmethod
    def validate_member_data(data: Dict, row_number: int) -> Tuple[bool, Optional[str]]:
        """
        Validate a single member record.

        Args:
            data: Dictionary containing member data
            row_number: Row number in the import file (for error reporting)

        Returns:
            Tuple of (is_valid, error_message)
        """
        # Check required fields
        for field in MemberImportValidator.REQUIRED_FIELDS:
            if not data.get(field) or str(data.get(field)).strip() == '':
                return False, f"Row {row_number}: Missing required field '{field}'"

        # Validate phone number format
        try:
            phone = str(data['phone_number']).strip()
            normalized_phone = normalize_phone_number(phone)
            data['phone_number'] = normalized_phone  # Update with normalized version
        except ValueError as e:
            return False, f"Row {row_number}: {str(e)}"

        # Validate email if provided
        email = data.get('email', '').strip()
        if email and '@' not in email:
            return False, f"Row {row_number}: Invalid email format"

        # Validate names are not too long
        first_name = str(data['first_name']).strip()
        last_name = str(data['last_name']).strip()

        if len(first_name) > 100:
            return False, f"Row {row_number}: First name too long (max 100 characters)"
        if len(last_name) > 100:
            return False, f"Row {row_number}: Last name too long (max 100 characters)"

        return True, None


class MemberImportParser:
    """
    Parses CSV and Excel files.
    Following SRP: Only responsible for file parsing.
    Following OCP: Can be extended for other formats.
    """

    @staticmethod
    def parse_csv(file_content: str) -> List[Dict]:
        """
        Parse CSV content into list of dictionaries.

        Args:
            file_content: CSV file content as string

        Returns:
            List of dictionaries, one per row
        """
        records = []
        csv_file = io.StringIO(file_content)
        reader = csv.DictReader(csv_file)

        for row in reader:
            # Clean up field names and values
            cleaned_row = {
                key.strip().lower(): value.strip() if value else ''
                for key, value in row.items()
                if key is not None
            }
            records.append(cleaned_row)

        return records

    @staticmethod
    def parse_excel(file_bytes: bytes) -> List[Dict]:
        """
        Parse Excel file into list of dictionaries.

        Args:
            file_bytes: Excel file content as bytes

        Returns:
            List of dictionaries, one per row
        """
        records = []
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True)
        sheet = workbook.active

        # Get header row
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).strip().lower() if cell.value else '')

        # Parse data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    record[headers[i]] = str(value).strip() if value else ''

            # Only add non-empty rows
            if any(record.values()):
                records.append(record)

        workbook.close()
        return records


class MemberImportService:
    """
    Orchestrates the member import process.
    Following SRP: Only responsible for coordinating import workflow.
    Following DIP: Depends on validator and parser abstractions.
    """

    def __init__(self):
        self.validator = MemberImportValidator()
        self.parser = MemberImportParser()

    def import_members(
        self,
        file_content: str,
        file_type: str = 'csv',
        batch_id: Optional[str] = None,
        send_notifications: bool = False
    ) -> Dict:
        """
        Import members from CSV or Excel file.

        Args:
            file_content: File content (string for CSV, base64 for Excel)
            file_type: 'csv' or 'excel'
            batch_id: Optional batch identifier for tracking
            send_notifications: Whether to send welcome SMS (default False)

        Returns:
            Dictionary with import results
        """
        # Parse file based on type
        try:
            if file_type == 'csv':
                records = self.parser.parse_csv(file_content)
            elif file_type == 'excel':
                import base64
                file_bytes = base64.b64decode(file_content)
                records = self.parser.parse_excel(file_bytes)
            else:
                return {
                    'success': False,
                    'message': f"Unsupported file type: {file_type}",
                    'imported_count': 0,
                    'skipped_count': 0,
                    'error_count': 0,
                    'errors': [],
                    'duplicates': []
                }
        except Exception as e:
            logger.error(f"Error parsing import file: {e}")
            return {
                'success': False,
                'message': f"Error parsing file: {str(e)}",
                'imported_count': 0,
                'skipped_count': 0,
                'error_count': 0,
                'errors': [str(e)],
                'duplicates': []
            }

        # Check record count limit
        if len(records) > MAX_IMPORT_RECORDS:
            return {
                'success': False,
                'message': f"File contains {len(records)} records. Maximum allowed is {MAX_IMPORT_RECORDS}.",
                'imported_count': 0,
                'skipped_count': 0,
                'error_count': 0,
                'errors': [f"Too many records ({len(records)}). Please split into smaller files of {MAX_IMPORT_RECORDS} or fewer."],
                'duplicates': []
            }

        if len(records) == 0:
            return {
                'success': False,
                'message': "No records found in file",
                'imported_count': 0,
                'skipped_count': 0,
                'error_count': 0,
                'errors': ["File contains no data rows"],
                'duplicates': []
            }

        # Generate batch ID if not provided
        if not batch_id:
            batch_id = f"import_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        # Process records
        return self._process_records(records, batch_id, send_notifications)

    def _process_records(
        self,
        records: List[Dict],
        batch_id: str,
        send_notifications: bool = False
    ) -> Dict:
        """
        Process and import member records in batches with transaction support.

        Args:
            records: List of member data dictionaries
            batch_id: Batch identifier
            send_notifications: Whether to send welcome SMS

        Returns:
            Import results dictionary
        """
        imported_count = 0
        skipped_count = 0
        error_count = 0
        errors = []
        duplicates = []

        # Get existing phone numbers and usernames for duplicate detection
        existing_phones = set(
            Member.objects.values_list('phone_number', flat=True)
        )
        existing_usernames = set(
            User.objects.values_list('username', flat=True)
        )

        # Validate all records first
        validated_records = []
        for i, record in enumerate(records, start=2):  # Start at 2 (row 1 is header)
            is_valid, error_msg = self.validator.validate_member_data(record, i)

            if not is_valid:
                errors.append(error_msg)
                error_count += 1
                continue

            # Check for duplicates
            phone = record['phone_number']
            if phone in existing_phones:
                duplicates.append(
                    f"Row {i}: Phone number {phone} already exists (skipped)"
                )
                skipped_count += 1
                continue

            # Add to existing phones to catch duplicates within the import file
            existing_phones.add(phone)
            validated_records.append(record)

        # Process validated records in batches
        all_imported_members = []
        batch_errors = []

        for batch_start in range(0, len(validated_records), BATCH_SIZE):
            batch = validated_records[batch_start:batch_start + BATCH_SIZE]
            batch_members = []

            try:
                with transaction.atomic():
                    for record in batch:
                        try:
                            member = self._create_member(record, batch_id)
                            # Also create a Django User for this member
                            user = self._create_user_for_member(
                                member, existing_usernames
                            )
                            if user:
                                member.user = user
                                member.save(update_fields=['user'])
                                existing_usernames.add(user.username)
                            batch_members.append(member)
                            imported_count += 1
                        except Exception as e:
                            row_num = batch_start + batch.index(record) + 2
                            batch_errors.append(
                                f"Row {row_num}: Error creating member - {str(e)}"
                            )
                            error_count += 1

                all_imported_members.extend(batch_members)

            except Exception as e:
                # Entire batch failed
                logger.error(f"Batch import failed (rows {batch_start+2}-{batch_start+len(batch)+1}): {e}")
                batch_errors.append(
                    f"Batch error (rows {batch_start+2}-{batch_start+len(batch)+1}): {str(e)}"
                )
                error_count += len(batch)

        errors.extend(batch_errors)

        # Optionally send welcome SMS (disabled by default)
        if send_notifications and all_imported_members:
            self._send_welcome_messages(all_imported_members)

        # Prepare response
        success = imported_count > 0
        message = self._generate_summary_message(
            imported_count, skipped_count, error_count
        )

        return {
            'success': success,
            'message': message,
            'imported_count': imported_count,
            'skipped_count': skipped_count,
            'error_count': error_count,
            'errors': errors,
            'duplicates': duplicates
        }

    def _create_member(self, record: Dict, batch_id: str) -> Member:
        """
        Create a member from validated record data.

        Args:
            record: Validated member data
            batch_id: Batch identifier

        Returns:
            Created Member instance
        """
        member = Member.objects.create(
            first_name=record['first_name'].strip(),
            last_name=record['last_name'].strip(),
            phone_number=record['phone_number'],
            email=record.get('email', '').strip() or None,
            member_number=record.get('member_number', '').strip() or None,
            is_active=True,
            is_guest=False,
            import_batch_id=batch_id
        )
        return member

    def _create_user_for_member(
        self,
        member: Member,
        existing_usernames: set
    ) -> Optional[User]:
        """
        Create a Django User account for an imported member so they can
        log in via OTP immediately.

        Args:
            member: The Member instance
            existing_usernames: Set of existing usernames to avoid duplicates

        Returns:
            Created User instance or None if user already exists
        """
        username = member.phone_number

        if username in existing_usernames:
            # User already exists, try to link
            try:
                existing_user = User.objects.get(username=username)
                return existing_user
            except User.DoesNotExist:
                pass

        try:
            user = User.objects.create(
                username=username,
                first_name=member.first_name,
                last_name=member.last_name,
                email=member.email or '',
                is_active=True,
            )
            # Set unusable password - auth is via OTP only
            user.set_unusable_password()
            user.save(update_fields=['password'])
            return user
        except Exception as e:
            logger.warning(f"Could not create user for member {member.phone_number}: {e}")
            return None

    def _send_welcome_messages(self, members: List[Member]):
        """Send welcome SMS to imported members. Failures don't block import."""
        try:
            from members.otp import SMSService
            sms_service = SMSService()

            sms_success = 0
            sms_failed = 0

            for member in members:
                try:
                    message = (
                        f"Welcome to SDA Church-Kawangware, {member.first_name}!\n\n"
                        f"Your member number is: {member.member_number}\n\n"
                        f"You can now make contributions via M-Pesa.\n\n"
                        f"God bless you!"
                    )
                    result = sms_service.send_sms(member.phone_number, message)
                    if result.get('success'):
                        sms_success += 1
                    else:
                        sms_failed += 1
                except Exception:
                    sms_failed += 1

            logger.info(f"Welcome SMS: {sms_success} sent, {sms_failed} failed")

        except Exception as e:
            logger.error(f"Error in SMS sending process: {e}")

    def _generate_summary_message(
        self,
        imported: int,
        skipped: int,
        errors: int
    ) -> str:
        """Generate human-readable summary message."""
        parts = []

        if imported > 0:
            parts.append(f"{imported} member{'s' if imported != 1 else ''} imported successfully")

        if skipped > 0:
            parts.append(f"{skipped} duplicate{'s' if skipped != 1 else ''} skipped")

        if errors > 0:
            parts.append(f"{errors} error{'s' if errors != 1 else ''}")

        if not parts:
            return "No members imported"

        return ", ".join(parts)

    @staticmethod
    def generate_csv_template() -> str:
        """
        Generate a CSV template for member import.

        Returns:
            CSV template as string
        """
        headers = ['first_name', 'last_name', 'phone_number', 'email']
        sample_data = [
            ['John', 'Doe', '0712345678', 'john@example.com'],
            ['Jane', 'Smith', '254723456789', 'jane@example.com'],
            ['Guest', 'Member', '0734567890', '']
        ]

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(sample_data)

        return output.getvalue()
