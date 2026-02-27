"""
Report Generation Services
Following SOLID principles:
- SRP: Each class has single responsibility
- OCP: Open for extension (can add new report types)
- DIP: Depends on abstractions (ReportGenerator, Exporter)
"""

from abc import ABC, abstractmethod
from typing import Dict, List, Optional
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum, Count, Q
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from contributions.models import Contribution, ContributionCategory
from members.models import Member


class ReportData:
    """Data container for report results"""
    def __init__(self, title: str, headers: List[str], rows: List[List], summary: Optional[Dict] = None):
        self.title = title
        self.headers = headers
        self.rows = rows
        self.summary = summary or {}
        self.generated_at = timezone.now()


class Exporter(ABC):
    """Abstract base class for exporters"""

    @abstractmethod
    def export(self, report_data: ReportData) -> BytesIO:
        """Export report data to file format"""
        pass


class ExcelExporter(Exporter):
    """
    Excel exporter using openpyxl.
    Following SRP: Only responsible for Excel export.
    """

    def export(self, report_data: ReportData) -> BytesIO:
        """Export report to Excel format"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Title
        ws.merge_cells('A1:' + chr(64 + len(report_data.headers)) + '1')
        title_cell = ws['A1']
        title_cell.value = report_data.title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Generated date
        ws.merge_cells('A2:' + chr(64 + len(report_data.headers)) + '2')
        date_cell = ws['A2']
        date_cell.value = f"Generated: {report_data.generated_at.strftime('%Y-%m-%d %H:%M')}"
        date_cell.alignment = Alignment(horizontal='center')

        # Headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(report_data.headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_num, row_data in enumerate(report_data.rows, 5):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Summary section
        if report_data.summary:
            summary_row = len(report_data.rows) + 6
            ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)

            for idx, (key, value) in enumerate(report_data.summary.items(), 1):
                ws.cell(row=summary_row + idx, column=1, value=key)
                ws.cell(row=summary_row + idx, column=2, value=str(value))

        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(report_data.headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            # Check all cells in this column
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


class PDFExporter(Exporter):
    """
    PDF exporter using ReportLab.
    Following SRP: Only responsible for PDF export.
    """

    def export(self, report_data: ReportData) -> BytesIO:
        """Export report to PDF format"""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            alignment=TA_CENTER,
            spaceAfter=12
        )

        # Title
        elements.append(Paragraph(report_data.title, title_style))
        elements.append(Paragraph(
            f"Generated: {report_data.generated_at.strftime('%Y-%m-%d %H:%M')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 0.3*inch))

        # Data table
        table_data = [report_data.headers] + report_data.rows

        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(table)

        # Summary
        if report_data.summary:
            elements.append(Spacer(1, 0.3*inch))
            elements.append(Paragraph("<b>Summary</b>", styles['Heading2']))

            summary_data = [[key, str(value)] for key, value in report_data.summary.items()]
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(summary_table)

        doc.build(elements)
        output.seek(0)
        return output


class ReportGenerator(ABC):
    """
    Abstract base class for report generators.
    Following OCP: Open for extension, closed for modification.
    """

    @abstractmethod
    def generate(self, **kwargs) -> ReportData:
        """Generate report data"""
        pass


class ContributionReportGenerator(ReportGenerator):
    """
    Generates contribution reports.
    Following SRP: Only responsible for contribution report generation.
    """

    def generate(
        self,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        category_id: Optional[int] = None,
        category_ids: Optional[List[int]] = None,
        member_id: Optional[int] = None,
        report_type: str = 'custom'
    ) -> ReportData:
        """
        Generate contribution report.

        Args:
            date_from: Start date
            date_to: End date
            category_id: Filter by single category (legacy)
            category_ids: Filter by multiple categories
            member_id: Filter by member
            report_type: Type of report (daily, weekly, monthly, custom)

        Returns:
            ReportData object
        """
        # Determine date range based on report type
        if report_type == 'daily':
            date_from = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = timezone.now()
            title = f"Daily Contribution Report - {date_from.strftime('%Y-%m-%d')}"
        elif report_type == 'weekly':
            today = timezone.now().date()
            date_from = datetime.combine(today - timedelta(days=today.weekday()), datetime.min.time())
            date_from = timezone.make_aware(date_from)
            date_to = timezone.now()
            title = f"Weekly Contribution Report - Week of {date_from.strftime('%Y-%m-%d')}"
        elif report_type == 'monthly':
            today = timezone.now()
            date_from = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            date_to = timezone.now()
            title = f"Monthly Contribution Report - {date_from.strftime('%B %Y')}"
        else:
            title = "Custom Contribution Report"
            if date_from and date_to:
                title += f" ({date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')})"

        # Build query
        queryset = Contribution.objects.select_related(
            'member', 'category', 'mpesa_transaction'
        ).filter(status='completed')

        if date_from:
            queryset = queryset.filter(transaction_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(transaction_date__lte=date_to)
        # Support both single category_id (legacy) and multiple category_ids
        if category_ids:
            queryset = queryset.filter(category_id__in=category_ids)
        elif category_id:
            queryset = queryset.filter(category_id=category_id)
        if member_id:
            queryset = queryset.filter(member_id=member_id)

        queryset = queryset.order_by('-transaction_date')

        # Build report data
        headers = ['Date', 'Member', 'Phone', 'Category', 'Amount (KES)', 'Receipt']
        rows = []

        total_amount = Decimal('0.00')

        for contribution in queryset:
            rows.append([
                contribution.transaction_date.strftime('%Y-%m-%d %H:%M') if contribution.transaction_date else 'N/A',
                contribution.member.full_name,
                contribution.member.phone_number,
                contribution.category.name,
                float(contribution.amount),
                contribution.mpesa_transaction.mpesa_receipt_number if contribution.mpesa_transaction else 'N/A'
            ])
            total_amount += contribution.amount

        # Summary
        summary = {
            'Total Contributions': queryset.count(),
            'Total Amount': f"KES {total_amount:,.2f}",
            'Average Amount': f"KES {(total_amount / queryset.count() if queryset.count() > 0 else 0):,.2f}"
        }

        return ReportData(title=title, headers=headers, rows=rows, summary=summary)


class ReportService:
    """
    Main service for report generation and export.
    Following DIP: Depends on abstractions (ReportGenerator, Exporter).
    """

    def __init__(self):
        self.contribution_generator = ContributionReportGenerator()
        self.excel_exporter = ExcelExporter()
        self.pdf_exporter = PDFExporter()

    def generate_contribution_report(
        self,
        format: str = 'excel',
        **kwargs
    ) -> BytesIO:
        """
        Generate and export contribution report.

        Args:
            format: Export format ('excel' or 'pdf')
            **kwargs: Arguments for report generation

        Returns:
            BytesIO object containing the exported report
        """
        # Generate report data
        report_data = self.contribution_generator.generate(**kwargs)

        # Export to requested format
        if format.lower() == 'pdf':
            return self.pdf_exporter.export(report_data)
        else:
            return self.excel_exporter.export(report_data)
